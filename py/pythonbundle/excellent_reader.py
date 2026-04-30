import openpyxl
import unidecode
from openpyxl.utils import get_column_letter, column_index_from_string

def _get_column_letter(column):
    if isinstance(column, int):
        return get_column_letter(column)
    return str(column).upper()

def _read_column(sheet, column_letter, skip_rows=2):
    column_data =[]
    counter = 0
    for cell in sheet[column_letter]:
        if counter >= skip_rows:
            val = str(cell.value).upper()
            column_data.append(val)
        counter += 1
    return column_data

def _write_column(sheet, column_letter, value, skip_rows=2):
    col_idx = column_index_from_string(column_letter)
    
    if isinstance(value, (list, tuple)):
        for i, val in enumerate(value):
            sheet.cell(row=skip_rows + 1 + i, column=col_idx, value=val)
    else:
        max_row = sheet.max_row
        start_row = skip_rows + 1
        
        if max_row < start_row:
            sheet.cell(row=start_row, column=col_idx, value=value)
        else:
            for row in range(start_row, max_row + 1):
                sheet.cell(row=row, column=col_idx, value=value)

def get_n_column_from_sheet_index(excel_file, index, column, skip_rows=2):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet_name = workbook.sheetnames[index]
        sheet = workbook[sheet_name]
        
        col_letter = _get_column_letter(column)
        data = {str(sheet_name).upper(): _read_column(sheet, col_letter, skip_rows)}
        return data
    except FileNotFoundError:
        print(f"Error: The file '{excel_file}' was not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def get_first_column_from_sheet_index(excel_file, index, skip_rows=2):
    return get_n_column_from_sheet_index(excel_file, index, 'A', skip_rows)

def get_n_column_from_all_sheets(excel_file, column, skip_rows=2):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        col_letter = _get_column_letter(column)
        data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data[str(sheet_name).upper()] = _read_column(sheet, col_letter, skip_rows)
        return data
    except FileNotFoundError:
        print(f"Error: The file '{excel_file}' was not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def get_first_column_from_all_sheets(excel_file, skip_rows=2):
    return get_n_column_from_all_sheets(excel_file, 'A', skip_rows)

def set_n_column_from_sheet_index(excel_file, index, column, value, skip_rows=2):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet_name = workbook.sheetnames[index]
        sheet = workbook[sheet_name]
        
        col_letter = _get_column_letter(column)
        _write_column(sheet, col_letter, value, skip_rows)
        
        workbook.save(excel_file)
        return True
    except FileNotFoundError:
        print(f"Error: The file '{excel_file}' was not found.")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False

def set_first_column_from_sheet_index(excel_file, index, value, skip_rows=2):
    return set_n_column_from_sheet_index(excel_file, index, 'A', value, skip_rows)

def set_n_column_from_all_sheets(excel_file, column, value, skip_rows=2):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        col_letter = _get_column_letter(column)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            _write_column(sheet, col_letter, value, skip_rows)
            
        workbook.save(excel_file)
        return True
    except FileNotFoundError:
        print(f"Error: The file '{excel_file}' was not found.")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False

def set_first_column_from_all_sheets(excel_file, value, skip_rows=2):
    return set_n_column_from_all_sheets(excel_file, 'A', value, skip_rows)